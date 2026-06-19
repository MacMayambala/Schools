from datetime import date
from decimal import Decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView
from django.contrib import messages
from django.contrib.auth.mixins import UserPassesTestMixin
from django.db import transaction, models
from django.db.models import Sum, F, Q, Value, DecimalField
from django.db.models.functions import Coalesce
from django.utils import timezone

from students.models import Student, Classroom
from .models import Account, Invoice, JournalEntry, JournalItem, Payment, FeeStructure, FeeLineItem, FeeCategory

from django.contrib.auth.mixins import UserPassesTestMixin

class FinanceDashboardView(UserPassesTestMixin, TemplateView):
    def test_func(self):
        # Only allow if the user's role has 'can_manage_finance' set to True
        return self.request.user.customuser.role.can_manage_finance


def finance_dashboard(request):
    school = request.school
    invoices = Invoice.objects.filter(school=school)

    # === Fee Collection Metrics ===
    total_expected = invoices.aggregate(
        total=Coalesce(Sum('current_fees'), Decimal('0'))
    )['total']

    total_collected = invoices.aggregate(
        total=Coalesce(Sum('paid_amount'), Decimal('0'))
    )['total']

    balance = total_expected - total_collected

    collection_rate = 0
    if total_expected > 0:
        collection_rate = round((total_collected / total_expected) * 100, 1)

    # === Chart of Accounts Based Financials ===
    total_income = Account.objects.filter(
        school=school,
        account_type='Income'
    ).aggregate(total=Coalesce(Sum('balance'), Decimal('0')))['total']

    total_expenses = Account.objects.filter(
        school=school,
        account_type='Expense'
    ).aggregate(total=Coalesce(Sum('balance'), Decimal('0')))['total']

    net_profit = total_income - total_expenses

    # === Student Ledger (FIXED QUERY) ===
    all_students = Student.objects.filter(
        school=school, 
        is_active=True
    ).select_related(
        'class_stream__classroom', # Changed from 'classroom' to follow the junction
        'class_stream__stream'     # Optional: added for better display
    ).annotate(
        total_billed=Coalesce(Sum('invoices__current_fees'), Decimal('0')),
        total_paid=Coalesce(Sum('invoices__paid_amount'), Decimal('0')),
        calc_balance=F('total_billed') - F('total_paid')
    ).order_by('first_name')

    context = {
        'total_expected': total_expected,
        'total_collected': total_collected,
        'balance': balance,
        'collection_rate': collection_rate,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'all_students': all_students,
        'today': timezone.now(),
        'recent_payments': Payment.objects.filter(school=school)
                            .select_related('invoice__student__class_stream__classroom') # Fixed path here too
                            .order_by('-date_paid')[:10],
    }

    return render(request, 'finance/dashboard.html', context)




    
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Invoice, FeeStructure
from students.models import Student, Classroom

from django.shortcuts import render, redirect
from django.contrib import messages
from .models import FeeStructure, Invoice
from students.models import Student, Classroom
#from .models import Classroom, FeeStructure, Student, Invoice  # Add your specific models here

from django.db import transaction
from django.contrib import messages
from django.shortcuts import render, redirect

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from students.models import Student, Classroom
from .models import Invoice, FeeStructure

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from django.db.models import Sum, F
from django.db.models.functions import Coalesce

from students.models import Student, Classroom
from .models import Invoice, FeeStructure


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from django.db.models import Sum, F
from django.db.models.functions import Coalesce

from students.models import Student, Classroom
from .models import Invoice, FeeStructure


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from django.db.models import Sum, F
from django.db.models.functions import Coalesce

from students.models import Student, Classroom
from .models import Invoice, FeeStructure


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from decimal import Decimal

from students.models import Student, Classroom
from .models import Invoice, FeeStructure
from django.db import transaction, models  # <--- Make sure 'models' is here

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, models
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from decimal import Decimal

from students.models import Student, Classroom
from .models import Invoice, FeeStructure

from django.db import transaction, models
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from .models import Invoice, FeeStructure
from students.models import Student, Classroom

from django.db import transaction, models
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from .models import Invoice, FeeStructure
from students.models import Student, Classroom

from django.db import transaction
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from decimal import Decimal
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from .models import Invoice, FeeStructure
from students.models import Student, Classroom


from django.db import transaction
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from decimal import Decimal
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from .models import Invoice, FeeStructure, FeeLineItem, FeeItem
from students.models import Student, Classroom


import datetime
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, models
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from datetime import date  # Change this to 'from datetime import date' for cleaner code

from .models import FeeStructure, Invoice, FeeLineItem, FeeCategory
from students.models import Student, Classroom

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from decimal import Decimal
from datetime import date
from .models import FeeStructure, Invoice, FeeLineItem, FeeCategory
from students.models import Student, Classroom

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction, models
from django.db.models import Sum, F
from django.db.models.functions import Coalesce
from decimal import Decimal
from datetime import date
from .models import FeeStructure, Invoice, FeeLineItem, FeeCategory
from students.models import Student, Classroom

from datetime import date
from decimal import Decimal

from django.contrib import messages
from django.db import transaction, models
from django.db.models import Sum, F, DecimalField
from django.db.models.functions import Coalesce
from django.shortcuts import render, redirect, get_object_or_404

from students.models import Student, Classroom
from .models import (
    Invoice,
    FeeStructure,
    FeeLineItem,
    FeeCategory
)


@transaction.atomic
def run_bulk_billing(request):
    if request.method == "POST":

        classroom_id = request.POST.get("classroom")
        term = request.POST.get("term")
        year = int(request.POST.get("year"))

        if not all([classroom_id, term, year]):
            messages.error(
                request,
                "Please select Class, Term and Year."
            )
            return redirect("finance:bulk_bill")

        classroom = get_object_or_404(
            Classroom,
            id=classroom_id,
            school=request.school
        )

        # =====================================
        # LOAD FEE STRUCTURES ONCE
        # =====================================

        structures = (
            FeeStructure.objects
            .filter(
                school=request.school,
                classroom=classroom,
                term=term,
                year=year
            )
            .prefetch_related(
                "template_items__category"
            )
        )

        if not structures.exists():
            messages.error(
                request,
                f"No fee structure defined for "
                f"{classroom.name} "
                f"Term {term} {year}"
            )
            return redirect("finance:bulk_bill")

        fee_map = {
            fs.section.strip().lower(): fs
            for fs in structures
        }

        # =====================================
        # TUITION CATEGORY
        # =====================================

        tuition_category, _ = FeeCategory.objects.get_or_create(
            school=request.school,
            name="Tuition",
            defaults={
                "is_mandatory": True
            }
        )

        # =====================================
        # STUDENTS
        # =====================================

        students = (
            Student.objects
            .filter(
                school=request.school,
                is_active=True,
                class_stream__classroom=classroom
            )
            .prefetch_related("invoices")
        )

        count = 0

        for student in students:

            section = (
                getattr(student, "section", None)
                or "Day"
            ).strip().lower()

            structure = fee_map.get(section)

            if not structure:
                continue

            # =====================================
            # ARREARS
            # =====================================

            previous_balance = (
                Invoice.objects
                .filter(
                    student=student,
                    school=request.school
                )
                .exclude(
                    term=term,
                    year=year
                )
                .aggregate(
                    balance=Coalesce(
                        Sum(
                            F("total_amount")
                            - F("paid_amount"),
                            output_field=DecimalField()
                        ),
                        Decimal("0.00")
                    )
                )["balance"]
            )

            # =====================================
            # INVOICE
            # =====================================

            invoice, created = Invoice.objects.get_or_create(
                student=student,
                school=request.school,
                term=term,
                year=year,
                defaults={
                    "current_fees": structure.total_fees,
                    "previous_balance": previous_balance,
                }
            )

            if not created:
                invoice.current_fees = structure.total_fees
                invoice.previous_balance = previous_balance
                invoice.save(
                    update_fields=[
                        "current_fees",
                        "previous_balance",
                        "total_amount",
                        "updated_at"
                    ]
                )

            # =====================================
            # REBUILD LINE ITEMS
            # =====================================

            invoice.invoice_items.all().delete()

            line_items = []

            # Tuition
            line_items.append(
                FeeLineItem(
                    school=request.school,
                    invoice=invoice,
                    category=tuition_category,
                    amount=structure.tuition_amount
                )
            )

            # Other charges
            for item in structure.template_items.all():
                line_items.append(
                    FeeLineItem(
                        school=request.school,
                        invoice=invoice,
                        category=item.category,
                        amount=item.amount
                    )
                )

            FeeLineItem.objects.bulk_create(
                line_items,
                batch_size=100
            )

            count += 1

        messages.success(
            request,
            f"{count} invoices generated successfully."
        )

        return redirect("finance:dashboard")

    classrooms = Classroom.objects.filter(
        school=request.school
    ).order_by("name")

    return render(
        request,
        "finance/bulk_bill.html",
        {
            "classrooms": classrooms,
            "current_year": date.today().year,
        },
    )

from django.shortcuts import render, redirect, get_object_or_404
from .models import Invoice, Payment
from .forms import PaymentForm

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Invoice, Payment
from .forms import PaymentForm
from .utils import send_payment_receipt_email  # Ensure this matches your file structure

# def record_payment(request, invoice_id):
#     # Fetch the invoice ensuring it belongs to the current school/tenant
#     invoice = get_object_or_404(Invoice, id=invoice_id, school=request.school)
    
#     if request.method == 'POST':
#         form = PaymentForm(request.POST)
#         if form.is_valid():
#             # 1. Prepare payment instance without saving to DB yet
#             payment = form.save(commit=False)
            
#             # 2. Assign relationships and Audit data
#             payment.invoice = invoice
#             payment.school = request.school
#             payment.recorded_by = request.user  # Captures the staff member 'serving' the user
            
#             # 3. Save to DB (triggers receipt_number generation & invoice total updates)
#             payment.save()
            
#             # 4. Trigger the branded email receipt
#             success = send_payment_receipt_email(request, payment)
            
#             # 5. Provide feedback to the staff member
#             if success:
#                 messages.success(request, f"Payment of {payment.amount_paid} recorded and receipt sent to parent.")
#             else:
#                 messages.warning(request, f"Payment of {payment.amount_paid} recorded, but receipt email failed to send.")
            
#             return redirect('finance:receipt_detail', payment_id=payment.id)
            
#     else:
#         # Default the payment amount to the remaining invoice balance
#         form = PaymentForm(initial={'amount_paid': invoice.balance})
        
#     return render(request, 'finance/record_payment.html', {
#         'form': form, 
#         'invoice': invoice
#     })

from django.shortcuts import render, get_object_or_404
from .models import Payment

def receipt_detail(request, payment_id):
    # select_related fetches the invoice and student in one query
    payment = get_object_or_404(
        Payment.objects.select_related('invoice__student__class_stream__classroom', 'school'), 
        id=payment_id, 
        school=request.school
    )
    return render(request, 'finance/receipt.html', {'payment': payment})


import logging
from django.contrib.sites.shortcuts import get_current_site
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.utils import timezone
from django.conf import settings

# Get an instance of a logger
logger = logging.getLogger(__name__)

def send_payment_receipt_email(request, payment):
    try:
        current_site = get_current_site(request)
        school = request.school 
        
        # 1. VALIDATE RECIPIENTS (The #1 reason emails fail)
        recipients = []
        
        # Check parent email
        if payment.invoice.student.parent_email:
            recipients.append(payment.invoice.student.parent_email)
        
        # Check school email
        if school.email:
            recipients.append(school.email)

        if not recipients:
            logger.warning(f"No recipients found for Payment {payment.receipt_number}. Email skipped.")
            return False

        # 2. PREPARE CONTENT
        subject = f"Receipt {payment.receipt_number} from {school.name}"
        context = {
            'payment': payment,
            'school': school,
            'protocol': 'https' if request.is_secure() else 'http',
            'domain': current_site.domain,
            'now': timezone.now(),
        }
        
        html_content = render_to_string('finance/emails/receipt_email.html', context)
        text_content = strip_tags(html_content)
        
        # 3. CONSTRUCT EMAIL
        # Format the 'From' address properly: "School Name <system@mail.com>"
        from_email = f"{school.name} <{settings.DEFAULT_FROM_EMAIL}>"
        
        email = EmailMultiAlternatives(
            subject, 
            text_content, 
            from_email, 
            recipients
        )
        email.attach_alternative(html_content, "text/html")
        
        # 4. SEND
        email.send(fail_silently=False)
        logger.info(f"Email sent successfully for {payment.receipt_number} to {recipients}")
        return True

    except Exception as e:
        # This will show up in your terminal/logs
        logger.error(f"Failed to send receipt email for {payment.receipt_number}: {str(e)}")
        return False

from django.db.models import F, Sum
from django.shortcuts import render
from .models import Invoice

from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import F, Sum
from django.db.models.functions import Coalesce

from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import F, Sum, DecimalField
from django.db.models.functions import Coalesce

from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import F, Sum, DecimalField
from django.db.models.functions import Coalesce

from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import F, Sum, DecimalField
from django.db.models.functions import Coalesce

from django.db.models import F, Sum, DecimalField
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.shortcuts import render

def defaulters_report(request):
    school = request.school
    classroom_id = request.GET.get('classroom')
    page_number = request.GET.get('page', 1)

    # Base queryset - Fixed select_related to use 'student__class_stream'
    defaulters_qs = Invoice.objects.filter(
        school=school,
        total_amount__gt=F('paid_amount')
    ).select_related('student', 'student__class_stream').order_by('-created_at')

    # Apply classroom filter if provided - Fixed to look up via class_stream_id
    if classroom_id:
        defaulters_qs = defaulters_qs.filter(student__class_stream_id=classroom_id)

    # Total count for badge and SMS modal (full records, not paginated)
    total_defaulters = defaulters_qs.count()

    # Calculate total outstanding balance
    total_debt = defaulters_qs.aggregate(
        total_billed=Coalesce(Sum('total_amount'), 0, output_field=DecimalField()),
        total_paid=Coalesce(Sum('paid_amount'), 0, output_field=DecimalField())
    )
    
    actual_balance = total_debt['total_billed'] - total_debt['total_paid']

    # Pagination
    paginator = Paginator(defaulters_qs, 15)
    page_obj = paginator.get_page(page_number)

    context = {
        'defaulters': page_obj,                 # Paginated records for table
        'total_defaulters': total_defaulters,   # Full count for badge & SMS
        'actual_balance': actual_balance,
        'classrooms': Classroom.objects.filter(school=school),
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
    }
    
    return render(request, 'finance/defaulters_list.html', context)

from decimal import Decimal
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Invoice, Payment, Account
from .utils import send_payment_receipt_email  # Import your email utility

from decimal import Decimal
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Invoice, Payment, Account
from .utils import send_payment_receipt_email

import uuid
from decimal import Decimal
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Invoice, Payment, Account
from .utils import send_payment_receipt_email

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from .models import Invoice, Payment, Account
from .utils import send_payment_receipt_email   # if you have this


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
import uuid
from .models import Invoice, Payment, Account
from .utils import send_payment_receipt_email


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
import uuid
from .models import Invoice, Payment, Account
from .utils import send_payment_receipt_email

import uuid
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from .models import Invoice, Account, Payment
# Ensure your utility import is correct
from .utils import send_payment_receipt_email 

def record_payment(request, invoice_id):
    # 1. Tenant Resolution
    active_school = getattr(request, 'school', None)
    
    if not active_school and request.user.is_authenticated:
        active_school = getattr(request.user, 'school', None) or \
                        Invoice.objects.filter(id=invoice_id).values_list('school', flat=True).first()

    # 2. Secure Invoice Retrieval
    invoice = Invoice.objects.filter(id=invoice_id, school=active_school).first()

    if not invoice:
        raw_exists = Invoice.objects.filter(id=invoice_id).exists()
        if raw_exists:
            messages.error(request, "Access Denied: You do not have permission to access this invoice.")
        else:
            messages.warning(request, "The requested invoice was not found.")
        return redirect('finance:dashboard')

    # 3. Handle Form Submission (POST)
    if request.method == "POST":
        try:
            with transaction.atomic():
                # Sanitize amount
                raw_amount = request.POST.get('amount', '0').replace(',', '')
                amount_val = Decimal(raw_amount)
                account_id = request.POST.get('account')

                if amount_val <= 0:
                    messages.error(request, "Please enter a valid payment amount greater than zero.")
                    return redirect('finance:record_payment', invoice_id=invoice.id)

                if amount_val > invoice.balance:
                    messages.error(request, f"Overpayment detected! Remaining balance is UGX {invoice.balance:,.0f}.")
                    return redirect('finance:record_payment', invoice_id=invoice.id)

                target_account = get_object_or_404(Account, id=account_id, school=active_school)

                # Create Payment
                payment = Payment.objects.create(
                    school=active_school,
                    invoice=invoice,
                    amount_paid=amount_val,
                    account=target_account,
                    payment_method=request.POST.get('method'),
                    transaction_id=request.POST.get('reference') or f"TX-{uuid.uuid4().hex[:8].upper()}",
                    depositor=request.POST.get('depositor'),
                    phone_number=request.POST.get('phone_number'),
                    recorded_by=request.user,
                    status='Completed'
                )

                # Update account balance
                target_account.balance += amount_val
                target_account.save()

            # Attempt to send receipt email (wrapped to prevent crash on mail error)
            try:
                send_payment_receipt_email(request, payment)
            except Exception as e:
                print(f"Receipt email failed: {e}")

            messages.success(request, f"Successfully recorded UGX {amount_val:,.0f} for {invoice.student.get_full_name()}.")
            
            # Redirect to the receipt detail after successful save
            return redirect('finance:receipt_detail', payment_id=payment.id)

        except Exception as e:
            messages.error(request, f"Transaction Failed: {str(e)}")
            return redirect('finance:record_payment', invoice_id=invoice.id)

    # 4. Handle Page View (GET)
    accounts = Account.objects.filter(
        school=active_school, 
        account_type__in=['Asset', 'Income']
    ).order_by('name')
    
    payments = Payment.objects.filter(invoice=invoice).order_by('-created_at')
    
    context = {
        'invoice': invoice,
        'accounts': accounts,
        'remaining_balance': invoice.balance,
        'student': invoice.student,
         
    }
    
    # This must be a RENDER, not a REDIRECT, to show the form
    return render(request, 'finance/record_payment.html', context)
   
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Account

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Account

def account_list(request):
    accounts = Account.objects.filter(school=request.school).order_by('account_type', 'code')

    if request.method == "POST":
        account_id = request.POST.get('account_id')  # For editing
        code = request.POST.get('code')
        name = request.POST.get('name')
        account_type = request.POST.get('account_type')
        opening_balance = request.POST.get('opening_balance', 0)

        try:
            if account_id:  # Edit existing account
                account = get_object_or_404(Account, id=account_id, school=request.school)
                # Check if code is being changed to one that already exists
                if code != account.code and Account.objects.filter(school=request.school, code=code).exists():
                    messages.error(request, f"Account code '{code}' already exists.")
                else:
                    account.code = code
                    account.name = name
                    account.account_type = account_type
                    account.balance = opening_balance
                    account.save()
                    messages.success(request, f"Account '{name}' updated successfully.")
            else:  # Create new account
                if Account.objects.filter(school=request.school, code=code).exists():
                    messages.error(request, f"Account code '{code}' already exists. Please use a unique code.")
                else:
                    Account.objects.create(
                        school=request.school,
                        code=code,
                        name=name,
                        account_type=account_type,
                        balance=opening_balance
                    )
                    messages.success(request, f"Account '{name}' created successfully.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

        return redirect('finance:account_list')

    context = {
        'accounts': accounts,
        'total_assets': accounts.filter(account_type='Asset').aggregate(Sum('balance'))['balance__sum'] or 0,
        'total_income': accounts.filter(account_type='Income').aggregate(Sum('balance'))['balance__sum'] or 0,
        'total_expenses': accounts.filter(account_type='Expense').aggregate(Sum('balance'))['balance__sum'] or 0,
    }
    return render(request, 'finance/account_list.html', context)


from django.utils import timezone
from django.db.models import Sum
from django.utils import timezone
from django.db.models import Sum

from django.utils import timezone

from django.utils import timezone
from django.db.models import Sum
from datetime import datetime

from django.db.models import Sum, F

from django.shortcuts import render
from django.utils import timezone
from django.db.models import Sum
from .models import Payment
# Ensure you have these imports if not already there

from django.db.models import Sum
from django.utils import timezone
from django.shortcuts import render

def daily_collection_report(request):
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    today = timezone.localtime(timezone.now()).date()
    
    # Use provided dates or default to today
    start_date = start_date_str if start_date_str else today
    end_date = end_date_str if end_date_str else today

    # Standardized query - ensure status matches your Payment model
    payments = Payment.objects.filter(
        school=request.school, 
        date_paid__date__range=[start_date, end_date],
        status='Completed'
    ).select_related(
        'invoice__student__class_stream__classroom', # Follow the relationship path
        'invoice__student'
    ).order_by('-date_paid')

    total_collected = payments.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    
    # Grouping by payment_method to match your model field
    summary_by_method = payments.values('payment_method').annotate(total=Sum('amount_paid'))

    return render(request, 'finance/daily_report.html', {
        'payments': payments,
        'total_collected': total_collected,
        'summary_by_method': summary_by_method,
        'start_date': start_date,
        'end_date': end_date,
        'today': today,
    })
import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from .models import Payment

def export_collections_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    payments = Payment.objects.filter(
        school=request.school,
        date_paid__date__range=[start_date, end_date]
    ).select_related('invoice__student', 'invoice__student__classroom', 'account')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collection_Report"

    headers = ['Date', 'Student Name', 'Class', 'Method', 'Account', 'Amount (UGX)', 'Receipt']
    ws.append(headers)

    for p in payments:
        ws.append([
            p.date_paid.strftime('%d/%m/%Y'),
            p.invoice.student.get_full_name(),
            p.invoice.student.classroom.name if p.invoice.student.classroom else "N/A",
            p.payment_method, # FIX: was p.method
            p.account.name if p.account else "N/A",
            p.amount_paid,    # FIX: was p.amount
            p.receipt_number
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Collections_{start_date}.xlsx'
    wb.save(response)
    return response
from django.db.models import Q

from django.db.models import Q
from django.shortcuts import render, get_object_or_404
from .models import Invoice, Payment

from django.shortcuts import render, get_object_or_404
from django.utils import timezone


from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from .models import Invoice, Payment


from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from .models import Invoice, Payment
from students.models import Student 
from django.shortcuts import render, get_object_or_404
from django.utils import timezone




def student_statement(request, student_id):
    student = get_object_or_404(Student, id=student_id, school=request.school)
    
    # Prefetch related data to avoid N+1 database queries
    invoices = Invoice.objects.filter(student=student, school=request.school).order_by('created_at')
    payments = Payment.objects.filter(invoice__student=student, school=request.school).order_by('date_paid')

    ledger = []

    # 1. Add Invoices (Debits) 
    for inv in invoices:
        # Use datetime for sorting, but store date for display
        dt = inv.created_at or timezone.now()
        ledger.append({
            'sort_key': dt, # Keep full timestamp for precise sorting
            'date': dt.date() if hasattr(dt, 'date') else dt, 
            'description': f"Fees Invoice - {inv.term} {inv.year}",
            'reference': getattr(inv, 'invoice_number', f"INV-{inv.id}"),
            'debit': inv.current_fees, 
            'credit': 0,
            'is_payment': False
        })
    
    # 2. Add Payments (Credits)
    for pay in payments:
        dt = pay.date_paid or timezone.now()
        # If date_paid is a date object, we might need to combine it with time for sorting
        sort_dt = timezone.make_aware(timezone.datetime.combine(dt, timezone.datetime.min.time())) if not hasattr(dt, 'hour') else dt

        # FIXED: Added () to execute get_payment_method_display method call
        if hasattr(pay, 'get_payment_method_display'):
            method_display = pay.get_payment_method_display()
        else:
            method_display = pay.payment_method

        ledger.append({
            'sort_key': sort_dt,
            'date': dt,
            'description': f"Fee Payment ({method_display})",
            'reference': pay.transaction_id or pay.receipt_number or f"PAY-{pay.id}",
            'debit': 0,
            'credit': pay.amount_paid,
            'is_payment': True
        })

    # 3. Sort by the full timestamp (sort_key)
    ledger.sort(key=lambda x: x['sort_key'])
    
    # 4. Calculate Running Balance
    running_total = 0
    for entry in ledger:
        running_total += (entry['debit'] - entry['credit'])
        entry['running_balance'] = running_total

    return render(request, 'finance/student_statement.html', {
        'student': student,
        'ledger': ledger,
        'final_balance': running_total,
        'today': timezone.now().date(),
    })

from django.shortcuts import render, redirect
from django.contrib import messages
from students.models import Classroom, Stream
from students.forms import ClassroomForm, StreamForm
from .models import FeeStructure
from .forms import FeeStructureForm

from django.shortcuts import render, redirect
from django.contrib import messages
from students.models import Classroom, Stream
from students.forms import ClassroomForm, StreamForm
from .models import FeeStructure
from .forms import FeeStructureForm
from django.shortcuts import render, redirect
from students.models import Classroom, Stream
from .models import FeeStructure

from django.shortcuts import render
from students.models import Classroom, Stream
from .models import FeeStructure
# Ensure your other imports (Forms) are here too
from django.shortcuts import render
from students.models import Classroom, Stream
from .models import FeeStructure
# Import your forms as well
def school_settings_hub(request):
    school = request.school
    
    if request.method == "POST":
        form_type = request.POST.get('form_type')

        if form_type == "classroom":
            form = ClassroomForm(request.POST)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.school = school
                obj.save()
                messages.success(request, f"Class {obj.name} added!")

        elif form_type == "stream":
            form = StreamForm(request.POST)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.school = school
                obj.save()
                messages.success(request, f"Stream {obj.name} added!")

        elif form_type == "fee_structure":
            form = FeeStructureForm(request.POST, school=school)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.school = school
                obj.save()
                messages.success(request, "Fee structure saved successfully!")
            else:
                messages.error(request, "Error saving fee structure.")

        return redirect('finance:settings_hub')

    # ✅ FIXED GET CONTEXT
    all_structs = FeeStructure.objects.filter(
        school=school
    ).select_related('classroom').prefetch_related('template_items__category')

    context = {
        'classrooms': Classroom.objects.filter(school=school),
        'streams': Stream.objects.filter(school=school),

        'fee_structures': all_structs,

        # 🔥 THESE WERE MISSING
        'term1_structures': all_structs.filter(term='1'),
        'term2_structures': all_structs.filter(term='2'),
        'term3_structures': all_structs.filter(term='3'),

        'class_form': ClassroomForm(),
        'stream_form': StreamForm(),
        'fee_form': FeeStructureForm(school=school),
    }

    return render(request, 'finance/settings_hub.html', context)
from django.shortcuts import render, get_object_or_404
from .models import Invoice

def print_invoice(request, invoice_id):
    invoice = get_object_or_404(Invoice, id=invoice_id, school=request.school)
    # Fetch payments linked to this invoice
    payments = invoice.payments.all()
    
    context = {
        'invoice': invoice,
        'payments': payments,
        'school': request.school, # From your multi-tenant middleware
    }
    return render(request, 'finance/print_invoice.html', context)



from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
import json

@csrf_exempt
def momo_webhook(request):
    """
    Listens for success/failure signals from the MoMo Provider.
    """
    if request.method == 'POST':
        data = json.loads(request.body)
        tx_ref = data.get('tx_ref')
        status = data.get('status') # e.g., 'successful'

        payment = Payment.objects.filter(transaction_id=tx_ref).first()
        
        if payment and status == 'successful':
            payment.status = 'Completed'
            payment.save() # This triggers Invoice.update_totals() automatically
            return HttpResponse(status=200)
            
    return HttpResponse(status=400)




from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Payment

def check_payment_status(request, payment_id):
    """
    Checks if a Mobile Money payment has been updated 
    to 'Completed' via the Webhook.
    """
    # Ensure we only check payments belonging to the current school/tenant
    payment = get_object_or_404(Payment, id=payment_id, school=request.school)
    
    return JsonResponse({
        'status': payment.status,
        'invoice_id': payment.invoice.id,
        'amount': float(payment.amount_paid),
        'balance': float(payment.invoice.balance)
    })


from students.models import Classroom, Stream
from students.forms import ClassroomForm, StreamForm # Assuming these exist
from .forms import FeeStructureForm
from students.forms import ClassroomForm, StreamForm


from django.db.models import Sum, F
from django.shortcuts import render
 # Adjust based on your actual model names

from django.db.models import Sum, F
from django.shortcuts import render
from .models import Invoice

from django.shortcuts import render
from django.db.models import Sum, Count, F, Q
from django.db.models.functions import Coalesce
from decimal import Decimal


def class_standing_report(request):
    # Filter by the current school (tenant)
    school = request.user.school 
    
    # We query Classrooms and pull student/invoice data through relationships
    report_data = Classroom.objects.filter(school=school).annotate(
        # Count unique students in this class
        student_count=Count('students', distinct=True),
        
        # Sum total_amount from all invoices belonging to students in this class
        total_expected=Coalesce(
            Sum('students__invoices__total_amount', filter=Q(students__school=school)), 
            Decimal('0')
        ),
        
        # Sum paid_amount from those same invoices
        total_collected=Coalesce(
            Sum('students__invoices__paid_amount', filter=Q(students__school=school)), 
            Decimal('0')
        )
    ).annotate(
        # Calculate balance and percentage in the database
        total_balance=F('total_expected') - F('total_collected')
    )

    # Calculate percentages in Python to handle division by zero safely
    for item in report_data:
        if item.total_expected > 0:
            item.collection_percentage = round((item.total_collected / item.total_expected) * 100, 1)
        else:
            item.collection_percentage = 0

    context = {
        'report_data': report_data,
        'currency': "UGX", # Or your preferred currency
    }
    return render(request, 'finance/class_standing.html', context)

from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .utils import send_bulk_fee_reminders

@login_required
def bulk_sms_reminder_view(request):
    """
    View to trigger bulk SMS reminders for all students with arrears.
    """
    # Optional: Get classroom_id from GET parameters if filtering by class
    classroom_id = request.GET.get('classroom_id')
    
    # Call the utility function
    sent, failed, status_msg = send_bulk_fee_reminders(request, classroom_id)
    
    if sent > 0:
        messages.success(request, f"Successfully sent {sent} fee reminders.")
    
    if failed > 0:
        # If some failed (usually due to balance or invalid phone numbers)
        messages.warning(request, f"Failed to send {failed} messages. Status: {status_msg}")
    
    if sent == 0 and failed == 0:
        messages.info(request, "No students with outstanding balances were found.")

    # Redirect back to the finance dashboard or student list
    return redirect(request.META.get('HTTP_REFERER', 'finance:dashboard'))

from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from weasyprint import HTML
from weasyprint.text.fonts import FontConfiguration   # ← This is the correct import
from django.template.loader import render_to_string

def student_report_card(request, student_id, term, year):
    student = get_object_or_404(Student, id=student_id, school=request.school)

    # ✅ Normalize
    clean_term = term.replace("Term-", "").strip()
    try:
        clean_year = int(year)
    except (ValueError, TypeError):
        clean_year = 2026

    # ✅ Fetch marks
    marks = Mark.objects.filter(
        student=student,
        term=clean_term,
        year=clean_year,
        school=request.school
    ).select_related('subject')

    # ✅ Calculate Average
    scores = [m.total_score for m in marks if m.total_score is not None]
    average = sum(scores) / len(scores) if scores else 0

    # ✅ Requirements Logic
    next_term_map = {"1": "2", "2": "3", "3": "1"}
    target_term = next_term_map.get(clean_term, "1")
    target_year = clean_year + 1 if clean_term == "3" else clean_year

    requirements = ClassRequirement.objects.filter(
        classroom=student.classroom,
        term=target_term,
        year=target_year,
        school=request.school
    ).first()

    # ✅ Rank Logic
    class_marks = Mark.objects.filter(
        classroom=student.classroom,
        term=clean_term,
        year=clean_year,
        school=request.school
    )
    totals = {}
    for m in class_marks:
        score_to_add = m.total_score if m.total_score is not None else 0
        totals[m.student_id] = totals.get(m.student_id, 0) + score_to_add

    sorted_totals = sorted(list(totals.values()), reverse=True)
    my_total = totals.get(student.id, 0)
    rank = sorted_totals.index(my_total) + 1 if my_total > 0 else "N/A"

    # ✅ Division Helper
    def get_division(student_marks):
        if not student_marks: return "N/A"
        grade_map = {'D1':1,'D2':2,'C3':3,'C4':4,'C5':5,'C6':6,'P7':7,'P8':8,'F9':9}
        points = [grade_map.get(m.grade, 9) for m in student_marks]
        points.sort()
        agg = sum(points[:8])
        if agg <= 12: return "Division 1"
        elif agg <= 24: return "Division 2"
        elif agg <= 34: return "Division 3"
        return "Division 4"

    # ✅ Comment Helper
    def get_teacher_comment(avg):
        if avg >= 75: return "An excellent result. Keep it up!"
        if avg >= 60: return "Good performance, but aim higher."
        if avg >= 45: return "Fairly good, more effort is needed."
        return "Poor performance. Please double your efforts."

    context = {
        'student': student,
        'marks': marks,
        'average': average,
        'rank': rank,
        'total_students': len(totals),
        'term': f"Term {clean_term}",
        'year': clean_year,
        'school': request.school,
        'division': get_division(marks),
        'teacher_comment': get_teacher_comment(average),
        'headteacher_comment': "Good progress. Keep it up.",
        'requirements': requirements,
    }

    # ====================== PDF GENERATION ======================
    html_string = render_to_string('academic/report_card.html', context)
    font_config = FontConfiguration()

    pdf = HTML(string=html_string).write_pdf(font_config=font_config)

    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f"Report_{student.admission_number}_{term}_{year}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response

# ====================== RECORD INCOME ======================
def record_income(request):
    if request.method == "POST":
        try:
            with transaction.atomic():
                account = get_object_or_404(
                    Account, 
                    id=request.POST.get('account'), 
                    school=request.school, 
                    account_type='Income'
                )
                amount = Decimal(request.POST.get('amount'))
                description = request.POST.get('description', 'Other Income')

                # Create Journal Entry
                journal = JournalEntry.objects.create(
                    school=request.school,
                    description=description,
                    date=timezone.now().date(),
                    created_by=request.user
                )

                # Credit Income Account
                JournalItem.objects.create(
                    journal=journal,
                    account=account,
                    credit=amount,
                    description=description,
                    school=request.school          # ← This was missing
                )

                # Debit Cash/Bank Account
                cash_account = Account.objects.filter(
                    school=request.school, 
                    account_type='Asset'
                ).first()
                
                if cash_account:
                    JournalItem.objects.create(
                        journal=journal,
                        account=cash_account,
                        debit=amount,
                        description=description,
                        school=request.school      # ← This was missing
                    )

                messages.success(request, f"Income of UGX {amount:,.0f} recorded successfully!")
                return redirect('finance:dashboard')

        except Exception as e:
            messages.error(request, f"Failed to record income: {str(e)}")

    income_accounts = Account.objects.filter(
        school=request.school, 
        account_type='Income', 
        is_active=True
    )

    return render(request, 'finance/record_income.html', {
        'income_accounts': income_accounts
    })


# ====================== RECORD EXPENSE ======================
def record_expense(request):
    if request.method == "POST":
        try:
            with transaction.atomic():
                account = get_object_or_404(
                    Account, 
                    id=request.POST.get('account'), 
                    school=request.school, 
                    account_type='Expense'
                )
                amount = Decimal(request.POST.get('amount'))
                description = request.POST.get('description', 'Expense')

                # Create Journal Entry
                journal = JournalEntry.objects.create(
                    school=request.school,
                    description=description,
                    date=timezone.now().date(),
                    created_by=request.user
                )

                # Debit Expense Account
                JournalItem.objects.create(
                    journal=journal,
                    account=account,
                    debit=amount,
                    description=description,
                    school=request.school           # ← This was missing
                )

                # Credit Cash/Bank Account
                cash_account = Account.objects.filter(
                    school=request.school, 
                    account_type='Asset'
                ).first()
                
                if cash_account:
                    JournalItem.objects.create(
                        journal=journal,
                        account=cash_account,
                        credit=amount,
                        description=description,
                        school=request.school       # ← This was missing
                    )

                messages.success(request, f"Expense of UGX {amount:,.0f} recorded successfully!")
                return redirect('finance:dashboard')

        except Exception as e:
            messages.error(request, f"Failed to record expense: {str(e)}")

    expense_accounts = Account.objects.filter(
        school=request.school, 
        account_type='Expense', 
        is_active=True
    )

    return render(request, 'finance/record_expense.html', {
        'expense_accounts': expense_accounts
    })


# finance/views/webhook.py
import hashlib
import json
import logging
from decimal import Decimal
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.utils import timezone
from django.conf import settings

from .models import Payment, Invoice, Account
from students.models import Student

logger = logging.getLogger(__name__)

def verify_schoolpay_signature(signature, receipt_number):
    api_password = getattr(settings, 'SCHOOLPAY_API_PASSWORD', None)
    if not api_password or not receipt_number:
        return False
    data = api_password + str(receipt_number)
    return hashlib.sha256(data.encode('utf-8')).hexdigest() == signature


@csrf_exempt
@transaction.atomic
def schoolpay_webhook(request):
    if request.method != 'POST':
        return HttpResponse(status=405)

    try:
        payload = json.loads(request.body)
        signature = payload.get('signature')
        payment_data = payload.get('payment', {})

        if not signature or not payment_data.get('schoolpayReceiptNumber'):
            return HttpResponse(status=400)

        receipt_number = payment_data['schoolpayReceiptNumber']
        amount = Decimal(payment_data.get('amount', 0))

        # Signature Check
        if not verify_schoolpay_signature(signature, receipt_number):
            logger.error(f"Invalid signature for receipt {receipt_number}")
            return HttpResponse(status=403)

        school = getattr(request, 'school', None)
        student_code = payment_data.get('studentPaymentCode')
        reg_number = payment_data.get('studentRegistrationNumber')

        # Find Student
        student = None
        if student_code:
            student = Student.objects.filter(studentPaymentCode=student_code).first()
        if not student and reg_number:
            student = Student.objects.filter(admission_number=reg_number).first()

        if not student:
            logger.warning(f"Student not found: {student_code or reg_number}")
            return HttpResponse(status=200)

        # Find or Create Invoice
        invoice = Invoice.objects.filter(student=student, school=school).order_by('-created_at').first()
        if not invoice:
            invoice = Invoice.objects.create(
                student=student,
                school=school,
                term="1",
                year=timezone.now().year,
                current_fees=amount,
                previous_balance=0,
                total_amount=amount,
            )

        # Get Account
        account = Account.objects.filter(school=school, account_type='Asset', is_active=True).first()

        # Create Payment
        Payment.objects.create(
            school=school,
            invoice=invoice,
            amount_paid=amount,
            payment_method='Mobile Money',
            receipt_number=receipt_number,
            transaction_id=payment_data.get('sourceChannelTransactionId') or receipt_number,
            depositor=payment_data.get('sourceChannelTransDetail', student.get_full_name()),
            phone_number=student.guardian_phone,
            status='Completed',
            account=account,
            recorded_by=None,   # Webhook payment
        )

        logger.info(f"✅ SchoolPay Success: {receipt_number} | UGX {amount} | Student: {student}")
        return HttpResponse(status=200)

    except Exception as e:
        logger.error(f"SchoolPay webhook failed: {str(e)}", exc_info=True)
        return HttpResponse(status=200)



########################################################################################################
import io
from decimal import Decimal
from django.shortcuts import render
from django.db.models import Sum, F, ExpressionWrapper, DecimalField, Q, Count
from django.db.models.functions import Cast, Coalesce
from django.http import HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.core.paginator import Paginator

# Export Imports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .models import Invoice
from students.models import Classroom, ClassStream

def get_financial_stats(request):
    school = request.user.school
    classroom_id = request.GET.get('classroom')
    sort_by = request.GET.get('sort', 'class_name')

    filters = Q(school=school)
    if classroom_id:
        filters &= Q(classroom_id=classroom_id)

    # Note: We use .select_related to avoid separate DB hits for class/stream names in the loop
    query = ClassStream.objects.filter(filters).select_related('classroom', 'stream', 'class_teacher').annotate(
        class_name=F('classroom__name'),
        stream_name=F('stream__name'),
        # This Count uses the 'related_name' we just set in the model
        total_students_count=Count('students', distinct=True),
        total_expected=Coalesce(
            Sum('students__invoices__total_amount'), 
            Value(0, output_field=DecimalField())
        ),
        total_paid=Coalesce(
            Sum('students__invoices__paid_amount'), 
            Value(0, output_field=DecimalField())
        ),
    )

    results = []
    grand_expected = Decimal('0.00')
    grand_paid = Decimal('0.00')

    for item in query:
        # We access the annotated fields directly
        expected = item.total_expected
        paid = item.total_paid
        balance = expected - paid
        
        # Safe division for percentage
        percent = 0
        if expected > 0:
            percent = round((float(paid) * 100 / float(expected)), 2)
        
        results.append({
            'class_name': item.class_name,
            'stream_name': item.stream_name,
            'total_students': item.total_students_count,
            'total_expected': expected,
            'total_paid': paid,
            'total_balance': balance,
            'pay_percent': percent,
            'teacher': item.class_teacher.get_full_name if item.class_teacher else "N/A"
        })
        grand_expected += expected
        grand_paid += paid
    # --- Sorting Logic ---
    if sort_by == 'highest_collection':
        results = sorted(results, key=lambda x: x['pay_percent'], reverse=True)
    elif sort_by == 'highest_balance':
        results = sorted(results, key=lambda x: x['total_balance'], reverse=True)
    else:
        results = sorted(results, key=lambda x: (x['class_name'], x['stream_name']))

    summary = {
        'grand_expected': grand_expected,
        'grand_paid': grand_paid,
        'grand_balance': grand_expected - grand_paid,
        'overall_percent': round((grand_paid * 100 / grand_expected), 2) if grand_expected > 0 else 0
    }

    return results, summary

def financial_report_view(request):
    results, summary = get_financial_stats(request)
    
    # Pagination
    paginator = Paginator(results, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'summary': summary,
        'classrooms': Classroom.objects.filter(school=request.user.school),
        'today': timezone.now(),
    }
    return render(request, 'finance/class_standing.html', context)

# --- EXPORT EXCEL ---
import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, PatternFill
from .models import Payment  # Ensure correct import based on your structure

def export_finance_excel(request):

    """
    Exports the payment collections report to Excel with a Total row at the bottom.
    """
    # 1. Fetch payments
    payments = Payment.objects.filter(school=request.school).select_related(
        'invoice__student__class_stream__classroom',
        'invoice__student__class_stream__stream',
        'account'
    ).order_by('-date_paid')

    # 2. Setup Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collection_Report"

    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    total_font = Font(bold=True)

    # 3. Headers
    headers = ['Date', 'Student Name', 'Class', 'Stream', 'Method', 'Account', 'Amount (UGX)', 'Receipt']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 4. Populate rows and track the total
    grand_total = 0
    for p in payments:
        student = p.invoice.student
        class_name = "N/A"
        stream_name = "N/A"

        if student.class_stream:
            class_name = student.class_stream.classroom.short_name
            stream_name = student.class_stream.stream.name
        
        amount = float(p.amount_paid)
        grand_total += amount

        ws.append([
            p.date_paid.strftime('%Y-%m-%d') if p.date_paid else "N/A",
            f"{student.first_name} {student.last_name}",
            class_name,
            stream_name,
            p.payment_method,
            p.account.name if p.account else "N/A",
            amount,
            p.receipt_number
        ])

    # 5. Add the Total Row
    ws.append([]) # Blank spacer row
    
    # We place "TOTAL" in the 'Account' column (6th) and the sum in 'Amount' column (7th)
    total_row_index = ws.max_row
    ws.cell(row=total_row_index, column=6, value="GRAND TOTAL:").font = total_font
    ws.cell(row=total_row_index, column=7, value=grand_total).font = total_font

    # 6. Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 3

    # 7. Return the file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Collection_Report_{timezone.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response
# --- EXPORT PDF ---
def export_finance_pdf(request):
    results, summary = get_financial_stats(request)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=1, spaceAfter=20)
    
    elements.append(Paragraph(f"{request.user.school.name} - Financial Report", title_style))
    
    data = [['Class', 'Stream', 'Students', 'Expected', 'Paid', 'Balance', '%']]
    for item in results:
        data.append([
            item['class_name'], item['stream_name'], item['total_students'],
            f"{item['total_expected']:,}", f"{item['total_paid']:,}", 
            f"{item['total_balance']:,}", f"{item['pay_percent']}%"
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Financial_Report.pdf'
    return response